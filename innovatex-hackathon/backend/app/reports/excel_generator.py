"""Excel report generator — openpyxl workbook.

Produces a workbook with 3 sheets:
  1. Resumen — metadata + global result + maturity
  2. Preguntas — per-question detail
  3. Recomendaciones — prioritized recommendations + action items
"""
from __future__ import annotations

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from app.schemas.report import ReportData

# ── Styling constants ──────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1E293B")
SUBTITLE_FONT = Font(name="Calibri", size=11, color="475569")
BODY_FONT = Font(name="Calibri", size=10, color="1E293B")
SCORE_FONT = Font(name="Calibri", size=24, bold=True, color="059669")
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
ALT_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
AMBER_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
GREEN_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")


def _style_header_row(ws, row: int, col_count: int):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_data_cell(ws, row: int, col: int, alt: bool = False):
    cell = ws.cell(row=row, column=col)
    cell.font = BODY_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    if alt:
        cell.fill = ALT_FILL


def generate_excel_bytes(data: ReportData) -> bytes:
    """Build an Excel workbook in memory and return the raw bytes."""
    wb = Workbook()

    # ── Sheet 1: Resumen ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"

    # Title
    ws1.merge_cells("A1:D1")
    ws1["A1"] = f"Diagnóstico de Cumplimiento — Ley 1581 de 2012"
    ws1["A1"].font = TITLE_FONT

    ws1.merge_cells("A2:D2")
    ws1["A2"] = f"{data.company_name}  ·  NIT {data.company_nit}  ·  {data.generated_at.strftime('%d/%m/%Y')}"
    ws1["A2"].font = SUBTITLE_FONT

    # Company info
    info_rows = [
        ("Empresa", data.company_name),
        ("NIT", data.company_nit),
        ("Sector", data.company_sector or "—"),
        ("Tamaño", data.company_size or "—"),
        ("Estado del diagnóstico", data.assessment_status),
        ("Completado el", data.completed_at.strftime("%d/%m/%Y") if data.completed_at else "—"),
    ]
    r = 4
    for label, value in info_rows:
        ws1.cell(row=r, column=1, value=label).font = Font(name="Calibri", size=10, bold=True)
        ws1.cell(row=r, column=2, value=value).font = BODY_FONT
        r += 1

    # Global score
    r += 1
    ws1.merge_cells(f"A{r}:B{r}")
    ws1.cell(row=r, column=1, value="Puntaje Global").font = Font(name="Calibri", size=12, bold=True)
    r += 1
    ws1.merge_cells(f"A{r}:B{r}")
    ws1.cell(row=r, column=1, value=f"{data.scores.overall_percentage:.1f}%").font = SCORE_FONT
    r += 1
    ws1.cell(row=r, column=1, value=data.scores.maturity_label).font = SUBTITLE_FONT

    # Block scores
    r += 2
    headers1 = ["Bloque", "Puntaje Obtenido", "Puntaje Máximo", "%"]
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=r, column=c, value=h)
    _style_header_row(ws1, r, len(headers1))
    for i, b in enumerate(data.scores.blocks):
        r += 1
        vals = [b.block_title, round(b.score, 2), round(b.max_score, 2), round(b.percentage, 1)]
        for c, v in enumerate(vals, 1):
            ws1.cell(row=r, column=c, value=v)
            _style_data_cell(ws1, r, c, alt=(i % 2 == 0))

    # Column widths
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 22
    ws1.column_dimensions["D"].width = 12

    # ── Sheet 2: Preguntas ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Preguntas")
    headers2 = ["#", "Bloque", "Pregunta", "Tipo", "Respuesta", "Notas"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
    _style_header_row(ws2, 1, len(headers2))

    for i, a in enumerate(data.answers):
        r = i + 2
        vals = [
            i + 1,
            a.block_slug,
            a.question_text,
            a.kind,
            a.answer_value or "—",
            a.notes or "",
        ]
        for c, v in enumerate(vals, 1):
            ws2.cell(row=r, column=c, value=v)
            _style_data_cell(ws2, r, c, alt=(i % 2 == 0))

    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 55
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 15
    ws2.column_dimensions["F"].width = 30

    # ── Sheet 3: Recomendaciones ───────────────────────────────────────────
    ws3 = wb.create_sheet("Recomendaciones")
    headers3 = ["Prioridad", "Recomendación", "Acción asociada", "Estado acción"]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=c, value=h)
    _style_header_row(ws3, 1, len(headers3))

    # Build a lookup: recommendation_id → action items
    ai_by_rec: dict[str, list] = {}
    for ai in data.action_items:
        # action_item has recommendation_id — but we don't have it in the read schema
        # Use title matching as a fallback
        pass

    # Map recommendations + paired action items
    row = 2
    recs = sorted(data.recommendations, key=lambda r: {"high": 0, "medium": 1, "low": 2}.get(r.priority.value, 3))
    for i, rec in enumerate(recs[:20]):
        prio = rec.priority.value
        prio_fill = {"high": RED_FILL, "medium": AMBER_FILL, "low": ALT_FILL}.get(prio, ALT_FILL)

        # Match action items by finding those that contain the recommendation snippet
        matched_ais = [ai for ai in data.action_items if rec.ai_generated_text[:50] in ai.title or ai.title in rec.ai_generated_text[:100]]
        if matched_ais:
            for ai in matched_ais:
                vals = [prio.upper(), rec.ai_generated_text, ai.title, ai.status]
                for c, v in enumerate(vals, 1):
                    ws3.cell(row=row, column=c, value=v)
                    _style_data_cell(ws3, row, c, alt=(i % 2 == 0))
                ws3.cell(row=row, column=1).fill = prio_fill
                row += 1
        else:
            vals = [prio.upper(), rec.ai_generated_text, "—", "—"]
            for c, v in enumerate(vals, 1):
                ws3.cell(row=row, column=c, value=v)
                _style_data_cell(ws3, row, c, alt=(i % 2 == 0))
            ws3.cell(row=row, column=1).fill = prio_fill
            row += 1

    # Unmatched action items
    if data.action_items and not data.recommendations:
        for i, ai in enumerate(data.action_items):
            vals = ["—", "—", ai.title, ai.status]
            for c, v in enumerate(vals, 1):
                ws3.cell(row=row, column=c, value=v)
                _style_data_cell(ws3, row, c, alt=(i % 2 == 0))
            row += 1

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 55
    ws3.column_dimensions["C"].width = 35
    ws3.column_dimensions["D"].width = 16

    # ── Write to buffer ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()
    buf.close()
    return excel_bytes
